"""
Test script to verify Excel file structure and access
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def test_excel_structure():
    file_path = Path(__file__).parent.parent / "Result_Sheet.xlsx"
    
    print("=" * 70)
    print("EXCEL FILE STRUCTURE TEST")
    print("=" * 70)
    print(f"\n📁 File Path: {file_path.absolute()}")
    print(f"✓ File Exists: {file_path.exists()}")
    
    if not file_path.exists():
        print("\n❌ Excel file not found!")
        return
    
    try:
        wb = load_workbook(file_path)
        print(f"\n✓ Workbook Loaded Successfully")
        print(f"📋 Available Sheets: {wb.sheetnames}")
        
        # Check BE sheet
        if 'BE' in wb.sheetnames:
            ws = wb['BE']
            print(f"\n{'='*70}")
            print("BE SHEET STRUCTURE")
            print("="*70)
            
            # Print Row 1 (Subject Headers)
            print("\n📊 Row 1 (Subject Headers):")
            row1_data = []
            for col in range(1, 20):
                value = ws.cell(row=1, column=col).value
                if value:
                    col_letter = get_column_letter(col)
                    row1_data.append(f"{col_letter}: {value}")
            print("   " + " | ".join(row1_data))
            
            # Print Row 2 (Sub-headers)
            print("\n📊 Row 2 (Sub-headers like CA1, MSE, CA2):")
            row2_data = []
            for col in range(1, 20):
                value = ws.cell(row=2, column=col).value
                if value:
                    col_letter = get_column_letter(col)
                    row2_data.append(f"{col_letter}: {value}")
            print("   " + " | ".join(row2_data))
            
            # Print first few student rows
            print("\n👥 First 5 Student Rows (starting from Row 3):")
            print("   Row | Col B (Roll No) | Col C (Name)")
            print("   " + "-"*60)
            for row in range(3, min(8, ws.max_row + 1)):
                roll_no = ws.cell(row=row, column=2).value
                name = ws.cell(row=row, column=3).value
                print(f"   {row:3d} | {str(roll_no):15s} | {str(name)}")
            
            # Find ZADE VISHAL SUNIL
            print("\n🔍 Searching for ZADE VISHAL SUNIL...")
            found = False
            for row in range(3, ws.max_row + 1):
                name = ws.cell(row=row, column=3).value
                if name and 'ZADE' in str(name).upper() and 'VISHAL' in str(name).upper():
                    roll_no = ws.cell(row=row, column=2).value
                    print(f"   ✓ Found at Row {row}")
                    print(f"   Roll Number: {roll_no}")
                    print(f"   Name: {name}")
                    found = True
                    
                    # Check DL columns
                    print("\n   📋 DL Subject Columns:")
                    for col in range(1, 20):
                        header = ws.cell(row=1, column=col).value
                        if header == 'DL':
                            col_letter = get_column_letter(col)
                            ca1_val = ws.cell(row=row, column=col).value
                            mse_val = ws.cell(row=row, column=col+1).value
                            ca2_val = ws.cell(row=row, column=col+2).value
                            print(f"      {col_letter}{row} (CA1): {ca1_val}")
                            print(f"      {get_column_letter(col+1)}{row} (MSE): {mse_val}")
                            print(f"      {get_column_letter(col+2)}{row} (CA2): {ca2_val}")
                    break
            
            if not found:
                print("   ❌ ZADE VISHAL SUNIL not found in BE sheet")
        else:
            print("\n❌ BE sheet not found in workbook!")
            
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_structure()
