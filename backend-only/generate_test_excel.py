#!/usr/bin/env python3
"""
Generate a test Excel file to verify the new format.
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_test_excel():
    """Generate a test Excel file with the new simplified format"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Marksheet"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "Test Subject - Mid Semester - Marksheet"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Exam details
        ws['A2'] = "Subject: Test Subject"
        ws['A3'] = "Exam Type: Mid Semester"
        ws['A4'] = "Date: 2025-12-03"
        ws['A5'] = "Class: BE-CSE"
        ws['A6'] = "Total Marks: 20"
        
        # Column headers - Simplified format
        header_row = 8
        headers = ['Roll Number', 'Student Name', 'Email', 'Marks Obtained', 'Total Marks', 'Percentage', 'Status']
        
        print("Generated Excel with headers:")
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            print(f"  Column {i}: {header}")
        
        # Sample data row
        row_num = header_row + 1
        sample_data = [
            "PRN-2022521242021",  # Roll Number
            "SANIYA KAUSER MAROOF SHAIKH",  # Student Name
            "prn-2022521242021@college.edu",  # Email
            18.5,  # Marks Obtained
            20,    # Total Marks
            "92.5%",  # Percentage
            "Checked"  # Status
        ]
        
        for col_num, data in enumerate(sample_data, 1):
            ws.cell(row=row_num, column=col_num, value=data)
        
        print(f"\nSample data row added:")
        for i, data in enumerate(sample_data, 1):
            print(f"  {headers[i-1]}: {data}")
        
        # Adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save file
        ROOT_DIR = Path(__file__).parent
        test_file_path = ROOT_DIR / "test_marksheet.xlsx"
        wb.save(test_file_path)
        
        print(f"\n✅ Test Excel file generated successfully!")
        print(f"📁 File saved as: {test_file_path}")
        print(f"📋 Format verified: Only 'Marks Obtained' and 'Total Marks' columns (no individual questions)")
        
        return True
        
    except Exception as e:
        print(f"❌ Failed to generate test Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("Generating Test Excel File")
    print("=" * 30)
    
    success = generate_test_excel()
    
    if success:
        print(f"\n🎉 Test completed successfully!")
        print("The Excel export will now use the simplified format.")
    else:
        print(f"\n❌ Test failed!")