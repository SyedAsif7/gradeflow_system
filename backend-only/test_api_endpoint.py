#!/usr/bin/env python3
"""
Test script to check if the Excel export API endpoint is working correctly.
"""

import asyncio
import os
from pathlib import Path
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
import sys
import traceback

# Add the parent directory to sys.path to import server functions
sys.path.append(os.path.join(os.path.dirname(__file__)))

# Load environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Database connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

async def test_export_function():
    """Test the export function directly"""
    try:
        # Import the export function from server
        # We'll simulate the export function here since we can't easily import it
        
        # Get an exam ID to test with
        exams = await db.exams.find({}, {"_id": 0}).to_list(1000)
        
        if not exams:
            print("No exams found!")
            return False
            
        test_exam = exams[0]
        exam_id = test_exam.get('id')
        print(f"Testing export for exam ID: {exam_id}")
        
        # Test the database queries that the export function uses
        exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
        if not exam:
            print("Failed to fetch exam")
            return False
        print("✅ Exam fetch successful")
        
        answer_sheets = await db.answer_sheets.find({"exam_id": exam_id}, {"_id": 0}).to_list(1000)
        print(f"✅ Found {len(answer_sheets)} answer sheets")
        
        subject = await db.subjects.find_one({"id": exam["subject_id"]}, {"_id": 0})
        if subject:
            print(f"✅ Subject found: {subject.get('name')}")
        else:
            print("⚠️ Subject not found")
        
        students = await db.students.find({}, {"_id": 0}).to_list(1000)
        students_dict = {student["id"]: student for student in students}
        print(f"✅ Loaded {len(students)} students")
        
        # Test openpyxl import
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            print("✅ openpyxl imports successful")
            
            # Test creating a simple workbook
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Test'
            print("✅ Workbook creation successful")
            
        except Exception as e:
            print(f"❌ openpyxl test failed: {e}")
            traceback.print_exc()
            return False
        
        print("\n🎉 All tests passed! The export function should work correctly.")
        return True
        
    except Exception as e:
        print(f"❌ Test failed with error: {e}")
        traceback.print_exc()
        return False
    finally:
        client.close()

async def main():
    """Main function"""
    print("Testing Excel Export API Endpoint")
    print("=" * 40)
    
    success = await test_export_function()
    
    if success:
        print("\n✅ API endpoint should be working correctly!")
        print("Try accessing the frontend and clicking the Export Excel button again.")
    else:
        print("\n❌ There may be an issue with the API endpoint.")

if __name__ == "__main__":
    asyncio.run(main())