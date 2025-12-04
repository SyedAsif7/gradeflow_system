from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import FileResponse, StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import shutil
import jwt
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from bson import ObjectId
from gridfs import NoFile

ROOT_DIR = Path(__file__).parent

# Load environment variables
load_dotenv(ROOT_DIR / '.env')
# Configure logging FIRST (before it's used)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Global variables for database connection
client = None
db = None
fs_bucket = None

# MongoDB connection with better error handling for serverless environments
def init_db():
    """Initialize database connection with better error handling"""
    global client, db, fs_bucket
    
    # Return early if already initialized
    if client and db:
        return client, db
    
    try:
        mongo_url = os.environ['MONGO_URL']
        db_name = os.environ['DB_NAME']
        
        # Log connection info (without exposing credentials)
        logger.info(f"Initializing MongoDB connection to database: {db_name}")
        
        # Create client
        client = AsyncIOMotorClient(
            mongo_url,
            maxPoolSize=10,
            minPoolSize=0,
            maxIdleTimeMS=60000,  # 60 seconds
            serverSelectionTimeoutMS=5000,  # 5 seconds
            connectTimeoutMS=5000,  # 5 seconds
        )
        
        db = client[db_name]
        logger.info("✅ MongoDB connection initialized successfully")
        
        # Initialize GridFS bucket
        fs_bucket = AsyncIOMotorGridFSBucket(db, bucket_name="answer_sheets")
        logger.info("✅ GridFS bucket initialized")
        
        return client, db
    except KeyError as e:
        logger.error(f"❌ Missing required environment variable: {e}")
        raise RuntimeError(f"Missing required environment variable: {e}. Please set MONGO_URL and DB_NAME.")
    except Exception as e:
        logger.error(f"❌ Failed to initialize MongoDB connection: {e}")
        raise RuntimeError(f"Failed to initialize MongoDB connection: {e}")

# Initialize database connection
try:
    client, db = init_db()
except Exception as e:
    logger.error(f"Failed to initialize database: {e}")
    # We'll re-raise this exception to prevent the app from starting with a broken DB
    raise

# Create upload directory
try:
    UPLOAD_DIR = ROOT_DIR / 'uploads' / 'answer_sheets'
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    logger.info(f"✅ Upload directory ready: {UPLOAD_DIR}")
except Exception as e:
    logger.error(f"❌ Failed to create upload directory: {e}")
    raise

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# JWT configuration
JWT_SECRET = os.environ.get("JWT_SECRET", "change_this_secret")
JWT_ALGORITHM = "HS256"
auth_scheme = HTTPBearer()

logger.info("✅ Server router initialization completed successfully")


def create_access_token(data: dict, expires_minutes: int = 60 * 24) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=expires_minutes)
    to_encode["exp"] = expire
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(auth_scheme),
):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
        # Get database connection lazily
        database = get_db()
        if database is None:
            raise HTTPException(status_code=503, detail="Database connection not available")
        user = await database.users.find_one({"email": email}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

def get_db():
    global client, db, fs_bucket
    if db is not None:
        return db
    try:
        client, db = init_db()
        return db
    except Exception:
        return None

def require_role(*roles: str):
    async def _dep(current_user=Depends(get_current_user)):
        if current_user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Forbidden")
        return current_user

    return _dep

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    role: str  # admin, teacher, student
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str

class LoginRequest(BaseModel):
    email: str
    password: str

class Student(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    roll_number: str
    class_name: str
    prn: Optional[int] = None
    semester: Optional[str] = None
    academic_year: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class StudentCreate(BaseModel):
    name: str
    email: str
    roll_number: str
    class_name: str
    prn: Optional[int] = None
    semester: Optional[str] = None
    academic_year: Optional[str] = None
    password: str

class Teacher(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    subject_ids: List[str] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TeacherCreate(BaseModel):
    name: str
    email: str
    subject_ids: List[str] = []
    password: str

class Subject(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    class_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SubjectCreate(BaseModel):
    name: str
    code: str
    class_name: Optional[str] = None

class Question(BaseModel):
    question_number: int
    question_text: str
    max_marks: int

class Exam(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: Optional[str] = None  # Deprecated - use exam_type instead
    subject_id: str
    exam_type: str  # CA-1, CA-2, Mid Semester
    date: str
    total_marks: int
    class_name: str
    questions: List[Question] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ExamCreate(BaseModel):
    name: Optional[str] = None  # Deprecated - use exam_type instead
    subject_id: str
    exam_type: str
    date: str
    total_marks: int
    class_name: str
    questions: List[Question] = []

class QuestionMark(BaseModel):
    question_number: int
    marks_obtained: float  # Changed from int to support half marks (0.5, 1.5, etc.)
    max_marks: int

class Annotation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str  # 'correct', 'incorrect', 'half-mark', 'quarter-mark', 'numeric', 'na', 'comment'
    question_number: int
    x: float  # Position on PDF (normalized 0-1)
    y: float  # Position on PDF (normalized 0-1)
    page: int
    value: Optional[str] = None  # For numeric marks or comment text
    marks: Optional[float] = None  # Marks associated with this annotation
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class AnswerSheet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    exam_id: str
    student_id: str
    pdf_filename: str
    assigned_teacher_id: Optional[str] = None
    status: str = "pending"  # pending, checked
    marks_obtained: Optional[float] = None  # Changed from int to support half marks
    question_marks: List[QuestionMark] = []
    annotations: List[Annotation] = []  # Store annotations for evaluation
    remarks: Optional[str] = None
    checked_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MarkSubmission(BaseModel):
    question_marks: Optional[List[QuestionMark]] = None
    total_marks: Optional[float] = None  # Changed from int to support half marks
    annotations: Optional[List[Annotation]] = None  # Annotations from evaluation interface
    remarks: Optional[str] = None

# Helper functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

async def persist_student_mark_to_excel(student: dict, subject: dict, exam: dict, marks_obtained: Optional[float]):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Use existing Excel file - DO NOT create new one
    file_path = Path(__file__).parent.parent / "Result_Sheet.xlsx"
    
    # If file doesn't exist, skip Excel export (don't create new file)
    if not file_path.exists():
        logger.info(f"Excel file not found at {file_path}. Skipping Excel export.")
        return

    # Load existing workbook
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return

    # Determine class name from exam or student
    class_name = exam.get('class_name', '') or student.get('class_name', 'BE')
    # Normalize to BE, TY, SY format
    if 'BTECH' in class_name.upper() or 'FINAL' in class_name.upper() or 'BE' in class_name.upper():
        sheet_name_base = 'BE'
    elif 'TY' in class_name.upper() or 'THIRD' in class_name.upper() or 'TE' in class_name.upper():
        sheet_name_base = 'TY'
    elif 'SY' in class_name.upper() or 'SECOND' in class_name.upper() or 'SE' in class_name.upper():
        sheet_name_base = 'SY'
    elif 'FY' in class_name.upper() or 'FIRST' in class_name.upper() or 'FE' in class_name.upper():
        sheet_name_base = 'FE'
    else:
        sheet_name_base = class_name[:31] if class_name else 'BE'  # Default to BE

    # Check if sheet exists - try variations with (1) suffix
    sheet_name = None
    possible_names = [
        f"{sheet_name_base}(1)",  # BE(1), TE(1), SE(1), FE(1)
        f"{sheet_name_base} (1)", # BE (1), TE (1), SE (1), FE (1) - with space
        sheet_name_base,           # BE, TE, SE, FE - without suffix
    ]
    
    for name in possible_names:
        if name in wb.sheetnames:
            sheet_name = name
            break
    
    if sheet_name is None:
        logger.warning(f"Sheet for class '{class_name}' not found in Excel file. Tried: {possible_names}. Available: {wb.sheetnames}")
        return
    
    ws = wb[sheet_name]

    # Get all headers from row 4 (not row 1!)
    headers = {}
    col_idx = 1
    while ws.cell(row=4, column=col_idx).value:
        headers[ws.cell(row=4, column=col_idx).value] = col_idx
        col_idx += 1
    max_col = col_idx - 1

    # Get subject code (short code preferred)
    subject_code = subject.get('code') or subject.get('name', 'Subject')
    
    # Map exam types to column names
    exam_type = exam.get('exam_type', '')
    if exam_type == 'CA-1':
        sub_col_name = 'CA1'
    elif exam_type == 'CA-2':
        sub_col_name = 'CA2'
    elif exam_type == 'Mid Semester':
        sub_col_name = 'MSE'
    else:
        sub_col_name = exam_type

    # Check if subject group exists in row 4
    subject_found = False
    subject_start_col = None
    
    # Look for subject in row 4 headers
    for col in range(4, max_col + 5):
        cell_value = ws.cell(row=4, column=col).value
        if cell_value == subject_code:
            subject_found = True
            subject_start_col = col
            break
    
    # If subject not found, skip export - DO NOT add new subjects
    if not subject_found:
        logger.warning(f"Subject '{subject_code}' not found in Excel sheet '{sheet_name}'. Available subjects in row 4: {list(headers.keys())}")
        return

    # Ensure row 5 has sub-headers (CA1, MSE, CA2)
    if ws.cell(row=5, column=subject_start_col).value != "CA1":
        ws.cell(row=5, column=subject_start_col, value="CA1")
        ws.cell(row=5, column=subject_start_col + 1, value="MSE")
        ws.cell(row=5, column=subject_start_col + 2, value="CA2")

    # Find student row by roll number - Students start from row 6
    roll_no = student.get('roll_number', 'N/A')
    student_name = student.get('name', 'Unknown')
    
    target_row = None
    for r in range(6, ws.max_row + 1):  # Changed from row 3 to row 6
        if ws.cell(row=r, column=2).value == roll_no:
            target_row = r
            break
    
    if target_row is None:
        logger.warning(f"Student '{student_name}' (Roll: {roll_no}) not found in Excel sheet '{sheet_name}'. Skipping export.")
        return

    # Determine which sub-column to write to
    if sub_col_name == 'CA1':
        target_col = subject_start_col
    elif sub_col_name == 'MSE':
        target_col = subject_start_col + 1
    elif sub_col_name == 'CA2':
        target_col = subject_start_col + 2
    else:
        target_col = subject_start_col + 1  # Default to MSE

    # Write marks
    ws.cell(row=target_row, column=target_col, value=marks_obtained if marks_obtained is not None else "")
    
    # Log the update for debugging
    col_letter = get_column_letter(target_col)
    logger.info(f"✅ Updated Excel: Student '{student_name}' (Roll: {roll_no}) | Subject: {subject_code} | Exam: {exam_type} | Cell: {col_letter}{target_row} | Marks: {marks_obtained}")

    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

    # Style the main headers
    header_font = Font(bold=True)
    for col in range(1, 4):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')

    wb.save(file_path)

# Auth routes
@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = user_data.model_dump()
    hashed_pwd = hash_password(user_dict.pop('password'))
    user_dict['password_hash'] = hashed_pwd
    
    user_obj = User(**{k: v for k, v in user_dict.items() if k != 'password_hash'})
    doc = user_obj.model_dump()
    doc['password_hash'] = hashed_pwd
    
    await db.users.insert_one(doc)
    return user_obj

@api_router.post("/auth/login")
async def login(login_data: LoginRequest):
    try:
        user = await db.users.find_one({"email": login_data.email}, {"_id": 0})
    except Exception as e:
        logger.error(f"MongoDB connection error: {str(e)}")
        raise HTTPException(
            status_code=503, 
            detail=f"Database connection error: {str(e)}. Please check MongoDB connection."
        )
    
    if not user:
        logger.warning(f"Login attempt with non-existent email: {login_data.email}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(login_data.password, user['password_hash']):
        logger.warning(f"Login attempt with wrong password for email: {login_data.email}")
        raise HTTPException(status_code=401, detail="Invalid credentials")

    user.pop('password_hash', None)
    token = create_access_token({"sub": user["email"], "role": user["role"]})
    return {"user": user, "token": token, "message": "Login successful"}

# Student routes
@api_router.post("/students", response_model=Student)
async def create_student(student_data: StudentCreate):
    existing = await db.students.find_one({"email": student_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    student_dict = student_data.model_dump()
    password = student_dict.pop('password')
    
    # Convert prn to int if provided
    if 'prn' in student_dict and student_dict['prn']:
        try:
            student_dict['prn'] = int(student_dict['prn'])
        except (ValueError, TypeError):
            student_dict['prn'] = None
    
    student_obj = Student(**student_dict)
    doc = student_obj.model_dump()
    
    await db.students.insert_one(doc)
    
    # Also create user account
    user_data = UserCreate(
        name=student_data.name,
        email=student_data.email,
        password=password,
        role="student"
    )
    await register(user_data)
    
    return student_obj

@api_router.get("/students", response_model=List[Student])
async def get_students(class_name: Optional[str] = None, semester: Optional[str] = None):
    """Get all students, optionally filtered by class_name (SY, TY, BE) or semester"""
    query = {}
    if class_name:
        query["class_name"] = class_name
    if semester:
        query["semester"] = semester
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    return students

@api_router.get("/students/{student_id}", response_model=Student)
async def get_student(student_id: str):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student

@api_router.put("/students/{student_id}", response_model=Student)
async def update_student(student_id: str, student_data: StudentCreate):
    existing = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Student not found")
    
    update_dict = student_data.model_dump(exclude={'password'})
    await db.students.update_one({"id": student_id}, {"$set": update_dict})
    
    updated = await db.students.find_one({"id": student_id}, {"_id": 0})
    return updated

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str):
    result = await db.students.delete_one({"id": student_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if student:
        await db.users.delete_one({"email": student.get('email')})
    
    return {"message": "Student deleted successfully"}

# Teacher routes
@api_router.post("/teachers", response_model=Teacher)
async def create_teacher(teacher_data: TeacherCreate):
    existing = await db.teachers.find_one({"email": teacher_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    teacher_dict = teacher_data.model_dump()
    password = teacher_dict.pop('password')
    
    teacher_obj = Teacher(**teacher_dict)
    doc = teacher_obj.model_dump()
    
    await db.teachers.insert_one(doc)
    
    user_data = UserCreate(
        name=teacher_data.name,
        email=teacher_data.email,
        password=password,
        role="teacher"
    )
    await register(user_data)
    
    return teacher_obj

@api_router.get("/teachers", response_model=List[Teacher])
async def get_teachers():
    teachers = await db.teachers.find({}, {"_id": 0}).to_list(1000)
    return teachers

@api_router.get("/teachers/{teacher_id}", response_model=Teacher)
async def get_teacher(teacher_id: str):
    teacher = await db.teachers.find_one({"id": teacher_id}, {"_id": 0})
    if not teacher:
        raise HTTPException(status_code=404, detail="Teacher not found")
    return teacher

@api_router.put("/teachers/{teacher_id}", response_model=Teacher)
async def update_teacher(teacher_id: str, teacher_data: TeacherCreate):
    existing = await db.teachers.find_one({"id": teacher_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Teacher not found")
    
    update_dict = teacher_data.model_dump(exclude={'password'})
    await db.teachers.update_one({"id": teacher_id}, {"$set": update_dict})
    
    updated = await db.teachers.find_one({"id": teacher_id}, {"_id": 0})
    return updated

@api_router.delete("/teachers/{teacher_id}")
async def delete_teacher(teacher_id: str):
    result = await db.teachers.delete_one({"id": teacher_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Teacher not found")
    
    teacher = await db.teachers.find_one({"id": teacher_id}, {"_id": 0})
    if teacher:
        await db.users.delete_one({"email": teacher.get('email')})
    
    return {"message": "Teacher deleted successfully"}

# Subject routes
@api_router.post("/subjects", response_model=Subject)
async def create_subject(subject_data: SubjectCreate):
    subject_obj = Subject(**subject_data.model_dump())
    doc = subject_obj.model_dump()
    await db.subjects.insert_one(doc)
    return subject_obj

@api_router.get("/subjects", response_model=List[Subject])
async def get_subjects(class_name: Optional[str] = None):
    """Retrieve subjects, optionally filtered by class/year."""
    query = {}
    if class_name:
        query["class_name"] = class_name
    subjects = await db.subjects.find(query, {"_id": 0}).to_list(1000)
    return subjects

@api_router.get("/subjects/{subject_id}", response_model=Subject)
async def get_subject(subject_id: str):
    subject = await db.subjects.find_one({"id": subject_id}, {"_id": 0})
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    return subject

@api_router.put("/subjects/{subject_id}", response_model=Subject)
async def update_subject(subject_id: str, subject_data: SubjectCreate):
    update_dict = subject_data.model_dump()
    result = await db.subjects.update_one({"id": subject_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    updated = await db.subjects.find_one({"id": subject_id}, {"_id": 0})
    return updated

@api_router.delete("/subjects/{subject_id}")
async def delete_subject(subject_id: str):
    result = await db.subjects.delete_one({"id": subject_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Subject not found")
    return {"message": "Subject deleted successfully"}

# Exam routes
@api_router.post("/exams", response_model=Exam)
async def create_exam(exam_data: ExamCreate):
    exam_obj = Exam(**exam_data.model_dump())
    doc = exam_obj.model_dump()
    await db.exams.insert_one(doc)
    return exam_obj

@api_router.get("/exams", response_model=List[Exam])
async def get_exams(credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))):
    # Try to get current user if token provided
    user = None
    if credentials:
        try:
            user = await get_current_user(credentials)
        except:
            pass  # If auth fails, continue without user
    
    # If teacher, filter exams by their assigned subjects
    if user and user.get("role") == "teacher":
        teacher = await db.teachers.find_one({"email": user["email"]}, {"_id": 0})
        if teacher and teacher.get("subject_ids"):
            exams = await db.exams.find(
                {"subject_id": {"$in": teacher["subject_ids"]}}, 
                {"_id": 0}
            ).to_list(1000)
            return exams
    
    # Admin or no auth - return all exams
    exams = await db.exams.find({}, {"_id": 0}).to_list(1000)
    return exams

@api_router.get("/exams/{exam_id}", response_model=Exam)
async def get_exam(exam_id: str):
    exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    return exam

@api_router.put("/exams/{exam_id}", response_model=Exam)
async def update_exam(exam_id: str, exam_data: ExamCreate):
    update_dict = exam_data.model_dump()
    result = await db.exams.update_one({"id": exam_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    updated = await db.exams.find_one({"id": exam_id}, {"_id": 0})
    return updated

@api_router.delete("/exams/{exam_id}")
async def delete_exam(exam_id: str):
    result = await db.exams.delete_one({"id": exam_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Exam not found")
    return {"message": "Exam deleted successfully"}

# Answer Sheet routes
@api_router.post("/answer-sheets/upload")
async def upload_answer_sheet(
    exam_id: str = Form(...),
    student_id: str = Form(...),
    assigned_teacher_id: str = Form(None),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    # Only allow admin and teacher roles to upload
    if current_user["role"] not in ["admin", "teacher"]:
        raise HTTPException(status_code=403, detail="Only admins and teachers can upload answer sheets")
    
    if not file.filename.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")

    # If teacher, validate they can upload for this exam's subject and auto-assign them
    if current_user["role"] == "teacher":
        # Get exam details
        exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
        if not exam:
            raise HTTPException(status_code=404, detail="Exam not found")
        
        # Get teacher details
        teacher = await db.teachers.find_one({"email": current_user["email"]}, {"_id": 0})
        if not teacher:
            raise HTTPException(status_code=404, detail="Teacher profile not found")
        
        # Check if teacher teaches this subject
        if exam["subject_id"] not in teacher.get("subject_ids", []):
            raise HTTPException(
                status_code=403, 
                detail="You can only upload answer sheets for exams related to your assigned subjects"
            )
        
        # Auto-assign the uploading teacher if no teacher was specified
        if not assigned_teacher_id:
            assigned_teacher_id = teacher["id"]

        # Prevent uploading if a checked sheet already exists for this exam and student
        existing_checked = await db.answer_sheets.find_one({
            "exam_id": exam_id,
            "student_id": student_id,
            "status": "checked",
        }, {"_id": 0})
        if existing_checked:
            raise HTTPException(status_code=403, detail="Upload blocked: paper already checked. Contact admin.")

    file_id = str(uuid.uuid4())
    metadata = {
        "original_name": file.filename,
        "content_type": file.content_type or "application/pdf",
        "exam_id": exam_id,
        "student_id": student_id,
        "assigned_teacher_id": assigned_teacher_id,
    }

    gridfs_id = await fs_bucket.upload_from_stream(file_id, file.file, metadata=metadata)

    answer_sheet = AnswerSheet(
        exam_id=exam_id,
        student_id=student_id,
        pdf_filename=str(gridfs_id),
        assigned_teacher_id=assigned_teacher_id,
    )
    doc = answer_sheet.model_dump()
    await db.answer_sheets.insert_one(doc)

    return answer_sheet

@api_router.get("/answer-sheets", response_model=List[AnswerSheet])
async def get_answer_sheets(
    teacher_id: Optional[str] = None, 
    student_id: Optional[str] = None,
    status: Optional[str] = None,
    limit: Optional[int] = 1000
):
    query = {}
    if teacher_id:
        query["assigned_teacher_id"] = teacher_id
    if student_id:
        query["student_id"] = student_id
    if status:
        query["status"] = status
    
    # Add projection to exclude heavy fields if not needed
    projection = {"_id": 0}
    
    answer_sheets = await db.answer_sheets.find(query, projection).limit(limit).to_list(limit)
    return answer_sheets

@api_router.get("/answer-sheets/{sheet_id}", response_model=AnswerSheet)
async def get_answer_sheet(sheet_id: str, mask_identity: bool = False):
    """
    Get answer sheet details.
    If mask_identity is True, student_id is still returned but frontend should mask it.
    """
    sheet = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Answer sheet not found")
    # Note: We still return student_id for backend operations, but frontend will mask it
    return sheet

@api_router.get("/answer-sheets/{sheet_id}/download")
async def download_answer_sheet(sheet_id: str):
    sheet = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Answer sheet not found")

    try:
        file_obj = await fs_bucket.open_download_stream(ObjectId(sheet["pdf_filename"]))
    except (NoFile, Exception):
        raise HTTPException(status_code=404, detail="File not found")

    async def file_iterator():
        # Increased chunk size for better performance
        while True:
            chunk = await file_obj.read(2 * 1024 * 1024)  # 2MB chunks
            if not chunk:
                break
            yield chunk

    filename = file_obj.metadata.get("original_name") if file_obj.metadata else "answer_sheet.pdf"

    return StreamingResponse(
        file_iterator(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Cache-Control": "public, max-age=3600",  # Cache for 1 hour
            "Accept-Ranges": "bytes",
        },
    )

@api_router.put("/answer-sheets/{sheet_id}/assign")
async def assign_answer_sheet(sheet_id: str, teacher_id: str = Form(...)):
    result = await db.answer_sheets.update_one(
        {"id": sheet_id},
        {"$set": {"assigned_teacher_id": teacher_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Answer sheet not found")
    
    updated = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    return updated

@api_router.put("/answer-sheets/{sheet_id}/grade", response_model=AnswerSheet)
async def grade_answer_sheet(sheet_id: str, marks_data: MarkSubmission):
    # Get exam to validate total marks
    sheet = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Answer sheet not found")
    
    exam = await db.exams.find_one({"id": sheet["exam_id"]}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Determine total marks: use direct input if provided, otherwise calculate from question-wise marks
    if marks_data.total_marks is not None:
        # Direct total marks input (for manual checking)
        if marks_data.total_marks < 0 or marks_data.total_marks > exam["total_marks"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Total marks must be between 0 and {exam['total_marks']}"
            )
        total_marks = marks_data.total_marks
        question_marks_list = marks_data.question_marks if marks_data.question_marks else []
    elif marks_data.question_marks and len(marks_data.question_marks) > 0:
        # Calculate from question-wise marks
        total_marks = sum(qm.marks_obtained for qm in marks_data.question_marks)
        question_marks_list = [qm.model_dump() for qm in marks_data.question_marks]
    else:
        raise HTTPException(
            status_code=400, 
            detail="Either total_marks or question_marks must be provided"
        )
    
    update_data = {
        "marks_obtained": total_marks,
        "question_marks": question_marks_list,
        "remarks": marks_data.remarks,
        "status": "checked",
        "checked_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Add annotations if provided
    if marks_data.annotations:
        update_data["annotations"] = [ann.model_dump() if isinstance(ann, Annotation) else ann for ann in marks_data.annotations]
    
    result = await db.answer_sheets.update_one(
        {"id": sheet_id},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Answer sheet not found")
    
    updated = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})

    # Persist marks to Excel sheet
    try:
        student = await db.students.find_one({"id": sheet["student_id"]}, {"_id": 0})
        subject = await db.subjects.find_one({"id": exam["subject_id"]}, {"_id": 0})
        
        if not student:
            logger.warning(f"Student not found for sheet {sheet_id}. Cannot persist to Excel.")
        elif not subject:
            logger.warning(f"Subject not found for exam {exam['id']}. Cannot persist to Excel.")
        else:
            await persist_student_mark_to_excel(student, subject, exam, update_data["marks_obtained"])
    except Exception as e:
        logger.error(f"Failed to persist marks to Excel: {str(e)}", exc_info=True)
        # Don't fail the API request if Excel update fails

    return updated

@api_router.delete("/answer-sheets/{sheet_id}")
async def delete_answer_sheet(sheet_id: str, current_user: dict = Depends(get_current_user)):
    # Only admin can delete answer sheets
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete answer sheets")
    sheet = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Answer sheet not found")

    # Try deleting the file from GridFS; ignore if it does not exist
    try:
        await fs_bucket.delete(ObjectId(sheet["pdf_filename"]))
    except Exception:
        pass

    await db.answer_sheets.delete_one({"id": sheet_id})
    return {"message": "Answer sheet deleted successfully"}

@api_router.put("/answer-sheets/{sheet_id}/reupload")
async def reupload_answer_sheet(sheet_id: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if not file.filename.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")

    sheet = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Answer sheet not found")

    # Disallow reupload for checked sheets for all roles
    if sheet.get("status") == "checked":
        raise HTTPException(status_code=403, detail="Reupload blocked: paper is already checked")

    # Teacher must be assigned to the subject of the exam
    if current_user["role"] == "teacher":
        exam = await db.exams.find_one({"id": sheet["exam_id"]}, {"_id": 0})
        if not exam:
            raise HTTPException(status_code=404, detail="Exam not found")
        teacher = await db.teachers.find_one({"email": current_user["email"]}, {"_id": 0})
        if not teacher:
            raise HTTPException(status_code=404, detail="Teacher profile not found")
        if exam["subject_id"] not in teacher.get("subject_ids", []):
            raise HTTPException(status_code=403, detail="Not allowed for non-assigned subjects")

    # Upload new PDF to GridFS
    file_id = str(uuid.uuid4())
    metadata = {
        "original_name": file.filename,
        "content_type": file.content_type or "application/pdf",
        "exam_id": sheet["exam_id"],
        "student_id": sheet["student_id"],
        "assigned_teacher_id": sheet.get("assigned_teacher_id"),
    }
    new_gridfs_id = await fs_bucket.upload_from_stream(file_id, file.file, metadata=metadata)

    # Delete old file if present
    try:
        await fs_bucket.delete(ObjectId(sheet["pdf_filename"]))
    except Exception:
        pass

    # Update the sheet with new file id
    await db.answer_sheets.update_one({"id": sheet_id}, {"$set": {"pdf_filename": str(new_gridfs_id)}})
    updated = await db.answer_sheets.find_one({"id": sheet_id}, {"_id": 0})
    return updated

# Dashboard stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    students_count = await db.students.count_documents({})
    teachers_count = await db.teachers.count_documents({})
    subjects_count = await db.subjects.count_documents({})
    exams_count = await db.exams.count_documents({})
    answer_sheets_count = await db.answer_sheets.count_documents({})
    pending_sheets = await db.answer_sheets.count_documents({"status": "pending"})
    
    return {
        "students": students_count,
        "teachers": teachers_count,
        "subjects": subjects_count,
        "exams": exams_count,
        "answer_sheets": answer_sheets_count,
        "pending_sheets": pending_sheets
    }

# Excel Export
@api_router.get("/exams/{exam_id}/export-marksheet")
async def export_marksheet(exam_id: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Fetch exam details
    exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Fetch subject
    subject = await db.subjects.find_one({"id": exam["subject_id"]}, {"_id": 0})
    subject_name = subject["name"] if subject else "Unknown"
    
    # Fetch all answer sheets for this exam
    answer_sheets = await db.answer_sheets.find({"exam_id": exam_id}, {"_id": 0}).to_list(1000)
    
    # Fetch all students
    students_dict = {}
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    for student in students:
        students_dict[student["id"]] = student
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Marksheet"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"{subject_name} - {exam['exam_type']} - Marksheet"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Exam details
    ws['A2'] = f"Subject: {subject_name}"
    ws['A3'] = f"Exam Type: {exam['exam_type']}"
    ws['A4'] = f"Date: {exam['date']}"
    ws['A5'] = f"Class: {exam['class_name']}"
    ws['A6'] = f"Total Marks: {exam['total_marks']}"
    
    # Column headers - Simplified to only show essential columns
    header_row = 8
    headers = ['Roll Number', 'Student Name', 'Email', 'Marks Obtained', 'Total Marks', 'Percentage', 'Status']
    
    # Note: Removed question-wise columns as requested
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row_num = header_row + 1
    for sheet in answer_sheets:
        student = students_dict.get(sheet["student_id"])
        if not student:
            continue
        
        col = 1
        ws.cell(row=row_num, column=col, value=student.get("roll_number", "N/A"))
        col += 1
        ws.cell(row=row_num, column=col, value=student.get("name", "N/A"))
        col += 1
        ws.cell(row=row_num, column=col, value=student.get("email", "N/A"))
        col += 1
        
        # Total marks obtained (no individual question marks as requested)
        marks_obtained = sheet.get("marks_obtained", 0)
        ws.cell(row=row_num, column=col, value=marks_obtained if marks_obtained is not None else "Not Graded")
        col += 1
        
        # Total marks
        ws.cell(row=row_num, column=col, value=exam["total_marks"])
        col += 1
        
        # Percentage
        if marks_obtained is not None:
            percentage = round((marks_obtained / exam["total_marks"]) * 100, 2)
            ws.cell(row=row_num, column=col, value=f"{percentage}%")
        else:
            ws.cell(row=row_num, column=col, value="N/A")
        col += 1
        
        # Status
        status = "Checked" if sheet["status"] == "checked" else "Pending"
        ws.cell(row=row_num, column=col, value=status)
        
        row_num += 1
    
    # Adjust column widths safely (skip merged cells)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in range(header_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # Also store to disk at fixed location
    try:
        from pathlib import Path
        result_path = Path(__file__).parent.parent / "Result_Sheet.xlsx"
        with open(result_path, "wb") as f:
            f.write(output.getbuffer())
    except Exception:
        pass
    
    # Generate filename - use exam_type if name is not available
    exam_name = exam.get('name') or exam.get('exam_type', 'Exam')
    filename = f"{exam_name.replace(' ', '_')}_Marksheet.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

@api_router.get("/admin/export-subject-results")
async def export_subject_results(class_name: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    # Load subjects (optionally filter by class/year)
    subj_query = {}
    if class_name:
        subj_query["class_name"] = class_name
    subjects = await db.subjects.find(subj_query, {"_id": 0}).to_list(1000)
    if not subjects:
        raise HTTPException(status_code=404, detail="No subjects found")

    # Preload teachers and students
    teachers = await db.teachers.find({}, {"_id": 0}).to_list(1000)
    teachers_by_id = {t["id"]: t for t in teachers}
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    students_by_id = {s["id"]: s for s in students}

    # Preload exams by subject id
    exams = await db.exams.find({}, {"_id": 0}).to_list(1000)
    exams_by_subject = {}
    for ex in exams:
        exams_by_subject.setdefault(ex["subject_id"], []).append(ex)

    # Workbook setup
    wb = Workbook()
    # Remove the default sheet; we'll add per-subject sheets
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for subject in subjects:
        ws = wb.create_sheet(title=(subject.get("name") or subject.get("code") or "Subject")[:31])
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Subject Results — {subject.get('name', subject.get('code', 'Unknown'))}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Column headers
        header_row = 3
        headers = [
            'Teacher', 'Exam Type', 'Class', 'Date',
            'Roll Number', 'Student Name', 'Marks Obtained', 'Total Marks', 'Status'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row_num = header_row + 1
        subject_exams = exams_by_subject.get(subject["id"], [])
        # For each exam under this subject, list all answer sheets
        for exam in subject_exams:
            sheets = await db.answer_sheets.find({"exam_id": exam["id"]}, {"_id": 0}).to_list(10000)
            for sheet in sheets:
                teacher = teachers_by_id.get(sheet.get("assigned_teacher_id"))
                student = students_by_id.get(sheet.get("student_id"))
                ws.cell(row=row_num, column=1, value=(teacher.get("name") if teacher else "Not Assigned"))
                ws.cell(row=row_num, column=2, value=exam.get("exam_type"))
                ws.cell(row=row_num, column=3, value=exam.get("class_name"))
                ws.cell(row=row_num, column=4, value=exam.get("date"))
                ws.cell(row=row_num, column=5, value=(student.get("roll_number") if student else "N/A"))
                ws.cell(row=row_num, column=6, value=(student.get("name") if student else "Unknown"))
                marks = sheet.get("marks_obtained")
                ws.cell(row=row_num, column=7, value=(marks if marks is not None else "Not Graded"))
                ws.cell(row=row_num, column=8, value=exam.get("total_marks"))
                status = "Checked" if sheet.get("status") == "checked" else "Pending"
                ws.cell(row=row_num, column=9, value=status)
                row_num += 1

        # Adjust column widths safely (skip merged cells)
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # Also store to disk at fixed location
    try:
        from pathlib import Path
        result_path = Path(__file__).parent.parent / "Result_Sheet.xlsx"
        with open(result_path, "wb") as f:
            f.write(output.getbuffer())
    except Exception:
        pass

    filename = "Subject_Wise_Teacher_Results.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )

app = FastAPI()

cors_origins = os.environ.get('CORS_ORIGINS', '*')
logger.info(f"CORS_ORIGINS config: {cors_origins}")
if cors_origins == '*':
    allow_origins = ['*']
else:
    allow_origins = [origin.strip() for origin in cors_origins.split(',')]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=allow_origins,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition", "Content-Type"],
)

@app.get("/")
async def root():
    return {
        "message": "GradeFlow System API",
        "version": "1.0.0",
        "docs": "/docs",
        "api_base": "/api",
    }

app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    if client:
        client.close()

__all__ = ['app', 'api_router', 'client', 'db', 'fs_bucket']
